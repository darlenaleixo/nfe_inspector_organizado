# -*- coding: utf-8 -*-
import os
import csv
import json
import logging
from typing import Dict, List, Any
import pandas as pd
from datetime import datetime
import io

# --- Dependências para o PDF Avançado ---
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

# --- Dependências para o Excel Avançado ---
try:
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# --- Dicionário para traduzir nomes de colunas no Excel ---
COLUNAS_TRADUZIDAS = {
    'status': 'Status', 'arquivo': 'Arquivo', 'chave_acesso': 'Chave de Acesso',
    'modelo_doc': 'Modelo', 'serie': 'Série', 'numero_nf': 'Número NF',
    'data_emissao': 'Data Emissão', 'valor_total_nf': 'Valor Total NF',
    'valor_total_produtos': 'Valor Total Produtos', 'emit_cnpj': 'CNPJ Emitente',
    'emit_nome': 'Nome Emitente', 'dest_cnpj_cpf': 'CNPJ/CPF Dest.', 'dest_nome': 'Nome Dest.',
    'pagamentos': 'Pagamentos', 'item_numero': 'Nº Item', 'item_codigo': 'Cód. Item',
    'item_descricao': 'Descrição Item', 'item_cfop': 'CFOP', 'item_quantidade': 'Qtd.',
    'item_valor_unitario': 'Vlr. Unitário', 'item_valor_total': 'Vlr. Total Item',
    'icms_cst': 'CST ICMS', 'icms_vbc': 'BC ICMS', 'icms_picms': 'Alíq. ICMS',
    'icms_vicms': 'Valor ICMS', 'icms_vbcst': 'BC ICMS-ST', 'icms_vicmsst': 'Valor ICMS-ST',
    'ipi_vipi': 'Valor IPI', 'pis_cst': 'CST PIS', 'pis_vbc': 'BC PIS',
    'pis_ppis': 'Alíq. PIS', 'pis_vpis': 'Valor PIS', 'cofins_cst': 'CST COFINS',
    'cofins_vbc': 'BC COFINS', 'cofins_pcofins': 'Alíq. COFINS', 'cofins_vcofins': 'Valor COFINS'
}

# --- FUNÇÃO PRINCIPAL ---
def gerar_todos_relatorios(pasta_saida: str, dados: List[Dict[str, Any]], resumos: Dict[str, Any], estatisticas: Dict[str, Any]):
    os.makedirs(pasta_saida, exist_ok=True)
    logging.info("Iniciando geração de todos os relatórios...")
    _gerar_csv(pasta_saida, dados)
    _gerar_json(pasta_saida, dados)
    _gerar_html(pasta_saida, resumos)
    if OPENPYXL_OK:
        _gerar_excel_formatado(pasta_saida, dados, resumos, estatisticas)
    else:
        logging.warning("Biblioteca 'openpyxl' não encontrada. O relatório Excel será gerado sem formatação avançada.")
        _gerar_excel_simples(pasta_saida, dados, resumos)
    if REPORTLAB_OK and MATPLOTLIB_OK:
        _gerar_pdf_avancado(pasta_saida, resumos, estatisticas)
    else:
        logging.warning("Bibliotecas para PDF avançado não encontradas (reportlab, matplotlib). Geração de PDF desabilitada.")

# --- Funções de Geração (CSV, JSON, HTML, Excel) ---
def _gerar_csv(pasta_saida, dados):
    caminho_csv = os.path.join(pasta_saida, "relatorio_detalhado.csv")
    if not dados: return
    try:
        chaves = sorted(list(set(k for d in dados for k in d.keys())))
        with open(caminho_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=chaves, extrasaction='ignore', delimiter=';')
            writer.writeheader()
            writer.writerows(dados)
        logging.info(f"Relatório CSV gerado: {caminho_csv}")
    except Exception as e:
        logging.error(f"Erro ao gerar CSV: {e}")

def _gerar_json(pasta_saida, dados):
    caminho_json = os.path.join(pasta_saida, "relatorio_detalhado.json")
    try:
        with open(caminho_json, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
        logging.info(f"Relatório JSON gerado: {caminho_json}")
    except Exception as e:
        logging.error(f"Erro ao gerar JSON: {e}")

def _gerar_html(pasta_saida, resumos):
    caminho_html = os.path.join(pasta_saida, "relatorio_dashboard.html")
    try:
        template_path = os.path.join(os.path.dirname(__file__), '..', 'ui', 'templates', 'dashboard.html')
        if not os.path.exists(template_path): raise FileNotFoundError
        with open(template_path, 'r', encoding='utf-8') as f:
            html = f.read()
        html = html.replace("$.getJSON('/api/resumos', function(data) {", f"var data = {json.dumps({'resumos': resumos})};")
        html = html.replace('serverSide: true, "ajax": "/api/dados",', 'serverSide: false,')
        with open(caminho_html, 'w', encoding='utf-8') as f:
            f.write(html)
        logging.info(f"Relatório HTML (dashboard estático) gerado: {caminho_html}")
    except Exception as e:
        logging.error(f"Erro ao gerar relatório HTML: {e}")

def _gerar_excel_simples(pasta_saida, dados, resumos):
    caminho_excel = os.path.join(pasta_saida, "relatorio_completo.xlsx")
    try:
        with pd.ExcelWriter(caminho_excel) as writer:
            pd.DataFrame(dados).to_excel(writer, sheet_name='Dados Completos', index=False)
            if resumos.get('apuracao_impostos'):
                pd.DataFrame(resumos['apuracao_impostos']).to_excel(writer, sheet_name='Apuracao por CFOP', index=True)
        logging.info(f"Relatório Excel (simples) gerado: {caminho_excel}")
    except Exception as e:
        logging.error(f"Erro ao gerar Excel simples: {e}")

def _gerar_excel_formatado(pasta_saida, dados, resumos, estatisticas):
    caminho_excel = os.path.join(pasta_saida, "relatorio_completo_formatado.xlsx")
    try:
        with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
            _criar_aba_dashboard_excel(writer, resumos, estatisticas)
            
            # Aba de Apuração por CFOP
            df_apuracao = pd.DataFrame(resumos.get('apuracao_impostos', {}))
            if not df_apuracao.empty:
                df_apuracao.index.name = 'CFOP'
                df_apuracao.reset_index(inplace=True)
                df_apuracao.to_excel(writer, sheet_name='Apuracao por CFOP', index=False, startrow=1)
                _formatar_aba_excel(writer.sheets['Apuracao por CFOP'], "Apuração de Impostos por CFOP")

            # --- NOVAS ABAS DE LIVROS FISCAIS ---
            df_entradas = pd.DataFrame.from_dict(resumos.get('livro_entradas', {}), orient='index')
            if not df_entradas.empty:
                df_entradas.index.name = 'CFOP'
                df_entradas.reset_index(inplace=True)
                df_entradas.to_excel(writer, sheet_name='Livro de Entradas', index=False, startrow=1)
                _formatar_aba_excel(writer.sheets['Livro de Entradas'], "Resumo de Operações de Entrada")

            df_saidas = pd.DataFrame.from_dict(resumos.get('livro_saidas', {}), orient='index')
            if not df_saidas.empty:
                df_saidas.index.name = 'CFOP'
                df_saidas.reset_index(inplace=True)
                df_saidas.to_excel(writer, sheet_name='Livro de Saídas', index=False, startrow=1)
                _formatar_aba_excel(writer.sheets['Livro de Saídas'], "Resumo de Operações de Saída")

            # Aba de Dados Completos (movida para o final para priorizar resumos)
            df_traduzido = pd.DataFrame(dados).rename(columns=COLUNAS_TRADUZIDAS)
            df_traduzido.to_excel(writer, sheet_name='Dados Completos', index=False, startrow=1)
            _formatar_aba_excel(writer.sheets['Dados Completos'], "Dados Detalhados das Notas Fiscais")

        logging.info(f"Relatório Excel formatado gerado: {caminho_excel}")
    except Exception as e:
        logging.error(f"Erro ao gerar Excel formatado: {e}")

# --- LÓGICA PARA O PDF AVANÇADO ---

class PageNumCanvas(object):
    def __init__(self, title): self.title = title
    def __call__(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9); canvas.setFillColor(colors.grey)
        canvas.drawString(2*cm, A4[1] - 1.5*cm, self.title)
        canvas.line(2*cm, A4[1] - 1.7*cm, A4[0] - 2*cm, A4[1] - 1.7*cm)
        canvas.setFont('Helvetica', 8); canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 2*cm, 1.5*cm, f"Página {doc.page}")
        canvas.drawString(2*cm, 1.5*cm, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        canvas.restoreState()

def _gerar_grafico_para_pdf(data: Dict, title: str, tipo='pie') -> io.BytesIO | None:
    if not data: return None
    buffer = io.BytesIO()
    plt.style.use('seaborn-v0_8-whitegrid')
    fig, ax = plt.subplots(figsize=(6, 2.5))
    if tipo == 'pie':
        ax.pie(data.values(), labels=data.keys(), autopct='%1.1f%%', startangle=140, textprops={'fontsize': 8})
        ax.axis('equal')
    elif tipo == 'barh':
        ax.barh(list(data.keys()), list(data.values()), color='skyblue')
        ax.invert_yaxis(); ax.tick_params(axis='y', labelsize=8)
    ax.set_title(title, fontsize=12, weight='bold')
    plt.tight_layout()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    plt.close(fig)
    buffer.seek(0)
    return buffer

def _formatar_valor_monetario(valor):
    try: return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError): return valor

def _criar_tabela_estilizada(data_list, col_widths=None):
    if not data_list: return None
    table = Table(data_list, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0056b3")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F0F8FF")),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    return table

def _gerar_pdf_avancado(pasta_saida, resumos, estatisticas):
    caminho_pdf = os.path.join(pasta_saida, "relatorio_analitico_profissional.pdf")
    doc = SimpleDocTemplate(caminho_pdf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2.5*cm, bottomMargin=2.5*cm)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('Title', parent=styles['h1'], alignment=TA_CENTER, fontSize=20, spaceAfter=1*cm)
    section_title_style = ParagraphStyle('SectionTitle', parent=styles['h2'], fontSize=16, spaceBefore=0.8*cm, spaceAfter=0.5*cm)
    
    story.append(Paragraph("Relatório Analítico de Notas Fiscais", title_style))
    story.append(Paragraph("Sumário Executivo (KPIs)", styles['h3']))
    kpi_data = [['Indicador', 'Valor']] + [[k, v] for k,v in {
        'Total de Vendas (Líquido)': _formatar_valor_monetario(resumos.get('total_vendas', 0)),
        'Total de Itens Vendidos': f"{resumos.get('total_itens_vendidos', 0):,}",
        'Notas Processadas': f"{estatisticas.get('notas_processadas_sucesso', 0):,}",
        'Notas Canceladas': f"{estatisticas.get('notas_canceladas', 0):,}",
    }.items()]
    story.append(_criar_tabela_estilizada(kpi_data))
    story.append(PageBreak())

    story.append(Paragraph("Análises Visuais e Operacionais", section_title_style))
    grafico_pagamentos = _gerar_grafico_para_pdf(resumos.get('formas_pagamento', {}), "Formas de Pagamento", tipo='pie')
    grafico_cfop = _gerar_grafico_para_pdf(resumos.get('top_cfops', {}), "Top CFOPs por Quantidade", tipo='barh')
    if grafico_pagamentos and grafico_cfop:
        story.append(Image(grafico_pagamentos, width=4*cm))
        story.append(Spacer(1, 0.5*cm))
        story.append(Image(grafico_cfop, width=4*cm))
    
    story.append(PageBreak())
    story.append(Paragraph("Apuração e Livros Fiscais", section_title_style))
    
    df_apuracao = pd.DataFrame(resumos.get('apuracao_impostos', {})).T.reset_index().rename(columns={'index':'CFOP'})
    if not df_apuracao.empty:
        story.append(Paragraph("Apuração de Impostos por CFOP", styles['h3']))
        for col in df_apuracao.columns[1:]: df_apuracao[col] = df_apuracao[col].apply(_formatar_valor_monetario)
        story.append(_criar_tabela_estilizada([df_apuracao.columns.tolist()] + df_apuracao.values.tolist()))
        story.append(Spacer(1, 0.5*cm))

    # --- NOVAS TABELAS DE LIVROS FISCAIS NO PDF ---
    livro_entradas = resumos.get('livro_entradas', {})
    if livro_entradas:
        story.append(Paragraph("Resumo de Operações de Entrada", styles['h3']))
        df_entradas = pd.DataFrame.from_dict(livro_entradas, orient='index').reset_index().rename(columns={'index':'CFOP'})
        for col in df_entradas.columns[1:]: df_entradas[col] = df_entradas[col].apply(_formatar_valor_monetario)
        story.append(_criar_tabela_estilizada([df_entradas.columns.tolist()] + df_entradas.values.tolist()))
        story.append(Spacer(1, 0.5*cm))

    livro_saidas = resumos.get('livro_saidas', {})
    if livro_saidas:
        story.append(Paragraph("Resumo de Operações de Saída", styles['h3']))
        df_saidas = pd.DataFrame.from_dict(livro_saidas, orient='index').reset_index().rename(columns={'index':'CFOP'})
        for col in df_saidas.columns[1:]: df_saidas[col] = df_saidas[col].apply(_formatar_valor_monetario)
        story.append(_criar_tabela_estilizada([df_saidas.columns.tolist()] + df_saidas.values.tolist()))

    try:
        canvas_handler = PageNumCanvas("Relatório Analítico de Notas Fiscais")
        doc.build(story, onFirstPage=canvas_handler, onLaterPages=canvas_handler)
        logging.info(f"Relatório PDF profissional gerado: {caminho_pdf}")
    except Exception as e:
        logging.error(f"Erro ao construir PDF: {e}")

# --- Funções Auxiliares para Excel ---
def _criar_aba_dashboard_excel(writer, resumos, estatisticas):
    ws = writer.book.create_sheet("Dashboard", 0)
    ws['B2'] = "Dashboard de Análise de Notas Fiscais"
    ws['B2'].font = Font(size=18, bold=True, color="000080")
    kpis = {
        "Total de Vendas": resumos.get('total_vendas', 0),
        "Total de Itens Vendidos": resumos.get('total_itens_vendidos', 0),
        "Notas Processadas": estatisticas.get('notas_processadas_sucesso', 0),
        "Notas Canceladas": estatisticas.get('notas_canceladas', 0),
        "Arquivos com Erro": estatisticas.get('arquivos_com_erro', 0),
        "Arquivos Inválidos (XSD)": estatisticas.get('arquivos_invalidos_xsd', 0)
    }
    row = 5
    for key, value in kpis.items():
        ws[f'B{row}'] = key; ws[f'C{row}'] = value
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        if "Vendas" in key: ws[f'C{row}'].number_format = 'R$ #,##0.00'
        row += 1
    ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 20

def _formatar_aba_excel(ws, titulo):
    ws.insert_rows(1); ws['A1'] = titulo
    ws['A1'].font = Font(size=16, bold=True, color="000080")
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    for col in ws.iter_cols(min_row=3):
        col_letter = col[0].column_letter
        header_text = str(ws[f'{col_letter}2'].value).lower()
        if any(term in header_text for term in ["valor", "icms", "ipi", "pis", "cofins"]):
            for cell in col:
                if isinstance(cell.value, (int, float)): cell.number_format = 'R$ #,##0.00'

