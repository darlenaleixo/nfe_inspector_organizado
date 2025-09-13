# -*- coding: utf-8 -*-

import os
import csv
import json
import logging
from typing import Dict, List, Any
import pandas as pd
from datetime import datetime
import tempfile

# --- Dependências ---
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Frame, PageTemplate, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import matplotlib.pyplot as plt
    import matplotlib.style as mplstyle
    mplstyle.use('seaborn-v0_8-whitegrid')
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

# --- Paleta de Cores Corporativa ---
PRIMARY_COLOR = colors.HexColor("#004C99")
SECONDARY_COLOR = colors.HexColor("#0066CC")
ACCENT_COLOR = colors.HexColor("#FF6B35")
LIGHT_BLUE = colors.HexColor("#E6F3FF")
LIGHT_GRAY = colors.HexColor("#F5F5F5")

# --- Dicionário para traduzir nomes de colunas ---
COLUNAS_TRADUZIDAS = {
    'status': 'Status', 'arquivo': 'Arquivo', 'chave_acesso': 'Chave de Acesso',
    'modelo_doc': 'Modelo', 'serie': 'Série', 'numero_nf': 'Número NF',
    'data_emissao': 'Data Emissão', 'valor_total_nf': 'Valor Total NF',
    'valor_total_produtos': 'Valor Total Produtos', 'emit_cnpj': 'CNPJ Emitente',
    'emit_nome': 'Nome Emitente', 'dest_cnpj_cpf': 'CNPJ/CPF Dest.', 'dest_nome': 'Nome Dest.',
    'pagamentos': 'Pagamentos', 'item_numero': 'Nº Item', 'item_codigo': 'Cód. Item',
    'item_descricao': 'Descrição Item', 'item_cfop': 'CFOP', 'item_ncm': 'NCM',
    'item_quantidade': 'Qtd.', 'item_valor_unitario': 'Vlr. Unitário', 'item_valor_total': 'Vlr. Total Item',
    'icms_cst': 'CST ICMS', 'icms_vbc': 'BC ICMS', 'icms_picms': 'Alíq. ICMS',
    'icms_vicms': 'Valor ICMS', 'icms_vbcst': 'BC ICMS-ST', 'icms_vicmsst': 'Valor ICMS-ST',
    'ipi_vipi': 'Valor IPI', 'pis_cst': 'CST PIS', 'pis_vbc': 'BC PIS',
    'pis_ppis': 'Alíq. PIS', 'pis_vpis': 'Valor PIS', 'cofins_cst': 'CST COFINS',
    'cofins_vbc': 'BC COFINS', 'cofins_pcofins': 'Alíq. COFINS', 'cofins_vcofins': 'Valor COFINS'
}

# --- FUNÇÃO PRINCIPAL ---
def gerar_todos_relatorios(pasta_saida: str, dados: List[Dict[str, Any]], resumos: Dict[str, Any], estatisticas: Dict[str, Any]):
    os.makedirs(pasta_saida, exist_ok=True)
    logging.info("Iniciando geração de todos os relatórios...")
    
    _gerar_csv(pasta_saida, dados)
    _gerar_json(pasta_saida, dados)
    
    if REPORTLAB_OK:
        _gerar_pdf_profissional(pasta_saida, dados, resumos, estatisticas)
    else:
        logging.warning("Biblioteca 'reportlab' não encontrada. Geração de PDF desabilitada.")
    
    if OPENPYXL_OK:
        _gerar_excel_formatado(pasta_saida, dados, resumos, estatisticas)
    else:
        logging.warning("Biblioteca 'openpyxl' não encontrada. Geração de Excel formatado desabilitada.")

# --- Funções de Geração (CSV, JSON) ---
def _gerar_csv(pasta_saida, dados):
    caminho_csv = os.path.join(pasta_saida, "relatorio_detalhado.csv")
    if not dados: return
    try:
        chaves = sorted(list(set(k for d in dados for k in d.keys())))
        with open(caminho_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=chaves, extrasaction='ignore', delimiter=';')
            writer.writeheader()
            writer.writerows(dados)
        logging.info(f"Relatório CSV gerado: {caminho_csv}")
    except Exception as e:
        logging.error(f"Erro ao gerar CSV: {e}")

def _gerar_json(pasta_saida, dados):
    caminho_json = os.path.join(pasta_saida, "relatorio_detalhado.json")
    try:
        with open(caminho_json, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
        logging.info(f"Relatório JSON gerado: {caminho_json}")
    except Exception as e:
        logging.error(f"Erro ao gerar JSON: {e}")

# --- FUNÇÕES AUXILIARES PARA PDF ---
def _formatar_valor_monetario(valor):
    try: 
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError): 
        return "R$ 0,00"

def _calcular_periodo_dados(dados):
    """Calcula o período dos dados baseado nas datas de emissão"""
    if not dados:
        return "Período não identificado"
    
    df = pd.DataFrame(dados)
    df['data_emissao'] = pd.to_datetime(df['data_emissao'], errors='coerce')
    df_validas = df.dropna(subset=['data_emissao'])
    
    if df_validas.empty:
        return "Período não identificado"
    
    data_min = df_validas['data_emissao'].min()
    data_max = df_validas['data_emissao'].max()
    
    if data_min.month == data_max.month and data_min.year == data_max.year:
        return f"{data_min.strftime('%B de %Y').title()}"
    else:
        return f"{data_min.strftime('%m/%Y')} a {data_max.strftime('%m/%Y')}"

def _gerar_grafico_vendas_por_mes(vendas_por_mes):
    """Gera gráfico de vendas por mês e retorna caminho do arquivo"""
    if not MATPLOTLIB_OK or not vendas_por_mes:
        return None
    
    try:
        fig, ax = plt.subplots(figsize=(12, 6))
        
        meses = list(vendas_por_mes.keys())
        valores = [float(v) for v in vendas_por_mes.values()]
        
        bars = ax.bar(meses, valores, color='#004C99', alpha=0.8)
        ax.set_title('Evolução de Vendas por Mês', fontsize=14, fontweight='bold')
        ax.set_ylabel('Valor (R$)', fontsize=12)
        ax.set_xlabel('Mês', fontsize=12)
        
        # Formatação dos valores no eixo Y
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'R$ {x:,.0f}'.replace(',', '.')))
        
        # Rotacionar labels do eixo X
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Salvar em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        plt.savefig(temp_file.name, dpi=300, bbox_inches='tight')
        plt.close()
        
        return temp_file.name
    except Exception as e:
        logging.error(f"Erro ao gerar gráfico: {e}")
        return None

def _gerar_grafico_formas_pagamento(formas_pagamento):
    """Gera gráfico de pizza para formas de pagamento"""
    if not MATPLOTLIB_OK or not formas_pagamento:
        return None
    
    try:
        fig, ax = plt.subplots(figsize=(10, 8))
        
        labels = list(formas_pagamento.keys())
        valores = [float(v) for v in formas_pagamento.values()]
        
        colors_pie = ['#004C99', '#0066CC', '#FF6B35', '#28A745', '#FFC107', '#DC3545']
        
        wedges, texts, autotexts = ax.pie(valores, labels=labels, autopct='%1.1f%%', 
                                         colors=colors_pie[:len(labels)], startangle=90)
        
        ax.set_title('Distribuição por Forma de Pagamento', fontsize=14, fontweight='bold')
        
        # Melhorar aparência dos textos
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        plt.savefig(temp_file.name, dpi=300, bbox_inches='tight')
        plt.close()
        
        return temp_file.name
    except Exception as e:
        logging.error(f"Erro ao gerar gráfico de pizza: {e}")
        return None

def _criar_tabela_profissional(data_list, col_widths=None, zebra=True):
    """Cria tabela com estilo profissional, incluindo zebra"""
    if not data_list: 
        return None
    
    table = Table(data_list, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
    
    # Estilo base
    ts = TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Corpo da tabela
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Alinha números à direita
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ])
    
    # Adicionar zebra nas linhas
    if zebra and len(data_list) > 1:
        for i in range(1, len(data_list)):
            bg_color = LIGHT_GRAY if i % 2 == 0 else colors.white
            ts.add('BACKGROUND', (0, i), (-1, i), bg_color)
    
    table.setStyle(ts)
    return table

class PageFooter(object):
    """Classe para rodapé profissional"""
    def __init__(self, emit_info, total_pages=None):
        self.emit_info = emit_info
        self.total_pages = total_pages

    def __call__(self, canvas, doc):
        canvas.saveState()
        w, h = doc.pagesize
        
        # Cabeçalho
        canvas.setFillColor(PRIMARY_COLOR)
        canvas.rect(0, h - 2*cm, w, 2*cm, fill=1)
        
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica-Bold', 12)
        canvas.drawString(2*cm, h - 1.2*cm, self.emit_info.get('nome', 'Empresa'))
        
        canvas.setFont('Helvetica', 10)
        canvas.drawString(2*cm, h - 1.6*cm, f"CNPJ: {self.emit_info.get('cnpj', 'N/A')}")
        
        # Rodapé
        canvas.setFillColor(colors.black)
        canvas.setFont('Helvetica', 8)
        canvas.drawString(2*cm, 1*cm, f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
        canvas.drawRightString(w - 2*cm, 1*cm, f"Página {doc.page}")
        
        canvas.restoreState()

def _gerar_pdf_profissional(pasta_saida, dados, resumos, estatisticas):
    """Gera PDF com layout profissional e melhorias visuais"""
    caminho_pdf = os.path.join(pasta_saida, "relatorio_gerencial_profissional.pdf")
    doc = SimpleDocTemplate(caminho_pdf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,    
                          topMargin=3*cm, bottomMargin=3*cm)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Informações da empresa
    info_emitente = {'nome': dados[0].get('emit_nome', 'Empresa'), 
                    'cnpj': dados[0].get('emit_cnpj', 'N/A')} if dados else {'nome': 'Empresa', 'cnpj': 'N/A'}
    
    periodo = _calcular_periodo_dados(dados)
    
    # Estilos personalizados
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], 
                                fontSize=24, textColor=PRIMARY_COLOR, alignment=TA_CENTER,
                                spaceBefore=2*cm, spaceAfter=1*cm, fontName='Helvetica-Bold')
    
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'],
                                   fontSize=16, textColor=SECONDARY_COLOR, alignment=TA_CENTER,
                                   spaceAfter=2*cm, fontName='Helvetica')
    
    section_style = ParagraphStyle('CustomSection', parent=styles['Heading2'],
                                  fontSize=16, textColor=PRIMARY_COLOR, fontName='Helvetica-Bold',
                                  spaceBefore=1*cm, spaceAfter=0.5*cm)
    
    # CAPA
    story.append(Spacer(1, 3*cm))
    story.append(Paragraph("RELATÓRIO GERENCIAL", title_style))
    story.append(Paragraph("OPERAÇÕES FISCAIS", title_style))
    story.append(Paragraph(periodo, subtitle_style))
    
    # Informações da empresa na capa
    empresa_style = ParagraphStyle('Empresa', parent=styles['Normal'], 
                                  fontSize=14, alignment=TA_CENTER, spaceAfter=0.3*cm)
    story.append(Paragraph(f"<b>{info_emitente['nome']}</b>", empresa_style))
    story.append(Paragraph(f"CNPJ: {info_emitente['cnpj']}", empresa_style))
    
    story.append(Spacer(1, 4*cm))
    story.append(Paragraph(f"Emitido em: {datetime.now().strftime('%d de %B de %Y')}", 
                          ParagraphStyle('DataEmissao', parent=styles['Normal'], 
                                       fontSize=12, alignment=TA_CENTER, textColor=colors.grey)))
    
    story.append(PageBreak())
    
    # SUMÁRIO EXECUTIVO
    story.append(Paragraph("SUMÁRIO EXECUTIVO", section_style))
    
    # KPIs principais com indicadores visuais
    kpi_data = [
        ['Indicador', 'Valor', 'Status'],
        ['Total de Vendas', _formatar_valor_monetario(resumos.get('total_vendas', 0)), '📈'],
        ['Total de Itens', f"{resumos.get('total_itens_vendidos', 0):,}".replace(',', '.'), '📦'],
        ['Notas Processadas', f"{estatisticas.get('notas_processadas_sucesso', 0):,}".replace(',', '.'), '📄'],
        ['Notas com Erro', f"{estatisticas.get('arquivos_com_erro', 0):,}".replace(',', '.'), '⚠️' if estatisticas.get('arquivos_com_erro', 0) > 0 else '✅']
    ]
    
    story.append(_criar_tabela_profissional(kpi_data, col_widths=[6*cm, 5*cm, 2*cm]))
    story.append(Spacer(1, 1*cm))
    
    # GRÁFICOS
    if resumos.get('vendas_por_mes') and len(resumos['vendas_por_mes']) > 1:
        grafico_vendas = _gerar_grafico_vendas_por_mes(resumos['vendas_por_mes'])
        if grafico_vendas:
            story.append(Paragraph("Evolução de Vendas", section_style))
            story.append(Image(grafico_vendas, width=16*cm, height=8*cm))
            story.append(Spacer(1, 1*cm))
    
    if resumos.get('formas_pagamento'):
        grafico_pagamentos = _gerar_grafico_formas_pagamento(resumos['formas_pagamento'])
        if grafico_pagamentos:
            story.append(PageBreak())
            story.append(Paragraph("Formas de Pagamento", section_style))
            story.append(Image(grafico_pagamentos, width=14*cm, height=10*cm))
            story.append(Spacer(1, 1*cm))
    
    # DETALHES FISCAIS
    story.append(PageBreak())
    story.append(Paragraph("APURAÇÃO FISCAL", section_style))
    
    # Filtrar CFOPs com valores > 0
    df_apuracao = pd.DataFrame(resumos.get('apuracao_impostos', {})).T
    if not df_apuracao.empty:
        # Filtrar apenas CFOPs com algum imposto > 0
        df_apuracao['total'] = df_apuracao.sum(axis=1)
        df_apuracao = df_apuracao[df_apuracao['total'] > 0].drop('total', axis=1)
        df_apuracao.reset_index(inplace=True)
        df_apuracao.rename(columns={'index': 'CFOP'}, inplace=True)
        
        if not df_apuracao.empty:
            story.append(Paragraph("Impostos por CFOP (apenas CFOPs com valores)", 
                                 ParagraphStyle('Subsection', parent=styles['Heading3'], 
                                              fontSize=12, spaceBefore=0.5*cm, spaceAfter=0.3*cm)))
            
            # Formatar valores monetários
            for col in df_apuracao.columns[1:]:
                df_apuracao[col] = df_apuracao[col].apply(_formatar_valor_monetario)
            
            data_apuracao = [df_apuracao.columns.tolist()] + df_apuracao.values.tolist()
            story.append(_criar_tabela_profissional(data_apuracao))
            story.append(Spacer(1, 1*cm))
    
    # ANOMALIAS E OBSERVAÇÕES
    auditoria = resumos.get('auditoria_aliquotas', [])
    if auditoria:
        story.append(PageBreak())
        story.append(Paragraph("ANOMALIAS FISCAIS", section_style))
        story.append(Paragraph("As seguintes divergências foram identificadas na análise das alíquotas de ICMS:",
                             styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        
        data_auditoria = [['Nota Fiscal', 'Produto', 'NCM', 'Alíq. Declarada', 'Alíq. Esperada', 'Diferença']]
        for item in auditoria:
            declarada = item.get('icms_picms', 0)
            esperada = item.get('aliquota_esperada', 0)
            diferenca = abs(declarada - esperada)
            data_auditoria.append([
                item.get('numero_nf', 'N/A'),
                item.get('item_descricao', 'N/A')[:30] + ('...' if len(item.get('item_descricao', '')) > 30 else ''),
                item.get('item_ncm', 'N/A'),
                f"{declarada:.2f}%",
                f"{esperada:.2f}%",
                f"{diferenca:.2f}%"
            ])
        
        story.append(_criar_tabela_profissional(data_auditoria))
        
        # Recomendações
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph("Recomendações:", ParagraphStyle('Bold', parent=styles['Normal'], fontName='Helvetica-Bold')))
        story.append(Paragraph("• Revisar as alíquotas de ICMS dos itens listados acima", styles['Normal']))
        story.append(Paragraph("• Verificar classificação fiscal (NCM) dos produtos com divergência", styles['Normal']))
        story.append(Paragraph("• Consultar legislação tributária específica para confirmação", styles['Normal']))
    
    # Configurar template de página
    footer = PageFooter(info_emitente)
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    template = PageTemplate(id='normal', frames=frame, onPage=footer)
    doc.addPageTemplates([template])
    
    try:
        doc.build(story)
        logging.info(f"Relatório PDF profissional gerado: {caminho_pdf}")
        
        # Limpar arquivos temporários de gráficos
        for temp_file in [_gerar_grafico_vendas_por_mes(resumos.get('vendas_por_mes', {})),
                         _gerar_grafico_formas_pagamento(resumos.get('formas_pagamento', {}))]:
            if temp_file and os.path.exists(temp_file):
                try:
                    os.unlink(temp_file)
                except:
                    pass
                    
    except Exception as e:
        logging.error(f"Erro ao construir o documento PDF: {e}")

# --- LÓGICA PARA O EXCEL (mantida igual) ---
def _gerar_excel_formatado(pasta_saida, dados, resumos, estatisticas):
    caminho_excel = os.path.join(pasta_saida, "relatorio_completo_formatado.xlsx")
    try:
        with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
            _criar_aba_dashboard_excel(writer, resumos, estatisticas)
            
            # Aba de Auditoria de ICMS
            auditoria_df = pd.DataFrame(resumos.get('auditoria_aliquotas', []))
            if not auditoria_df.empty:
                auditoria_df.to_excel(writer, sheet_name='Auditoria ICMS', index=False, startrow=1)
                _formatar_aba_excel(writer.sheets['Auditoria ICMS'], "Divergências de Alíquotas de ICMS")
            
            # Outras abas...
            df_apuracao = pd.DataFrame(resumos.get('apuracao_impostos', {}))
            if not df_apuracao.empty:
                df_apuracao.index.name = 'CFOP'
                df_apuracao.reset_index(inplace=True)
                df_apuracao.to_excel(writer, sheet_name='Apuracao por CFOP', index=False, startrow=1)
                _formatar_aba_excel(writer.sheets['Apuracao por CFOP'], "Apuração de Impostos por CFOP")
            
            df_traduzido = pd.DataFrame(dados).rename(columns=COLUNAS_TRADUZIDAS)
            df_traduzido.to_excel(writer, sheet_name='Dados Completos', index=False, startrow=1)
            _formatar_aba_excel(writer.sheets['Dados Completos'], "Dados Detalhados das Notas Fiscais")
        
        logging.info(f"Relatório Excel formatado gerado: {caminho_excel}")
    except Exception as e:
        logging.error(f"Erro ao gerar Excel formatado: {e}")

def _criar_aba_dashboard_excel(writer, resumos, estatisticas):
    ws = writer.book.create_sheet("Dashboard", 0)
    ws['B2'] = "Dashboard de Análise de Notas Fiscais"
    ws['B2'].font = Font(size=18, bold=True, color="000080")
    
    kpis = {
        "Total de Vendas": resumos.get('total_vendas', 0),
        "Total de Itens Vendidos": resumos.get('total_itens_vendidos', 0),
        "Notas Processadas": estatisticas.get('notas_processadas_sucesso', 0),
    }
    
    row = 5
    for key, value in kpis.items():
        ws[f'B{row}'] = key; ws[f'C{row}'] = value
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        if "Vendas" in key: ws[f'C{row}'].number_format = 'R$ #,##0.00'
        row += 1
    
    ws.column_dimensions['B'].width = 30; ws.column_dimensions['C'].width = 20

def _formatar_aba_excel(ws, titulo):
    ws.insert_rows(1); ws['A1'] = titulo
    ws['A1'].font = Font(size=16, bold=True, color="000080")
    
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    for col in ws.iter_cols(min_row=3):
        col_letter = col[0].column_letter
        header_text = str(ws[f'{col_letter}2'].value).lower()
        if any(term in header_text for term in ["valor", "icms", "ipi", "pis", "cofins"]):
            for cell in col:
                if isinstance(cell.value, (int, float)): cell.number_format = 'R$ #,##0.00'
                